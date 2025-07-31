import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import io
import asyncio
from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from main import app, create_workbook, parse_excel, jobs


def make_excel(voc=1.1, jsc=2.2, ff=0.9) -> bytes:
    df = pd.DataFrame({"Voc": [voc], "Jsc": [jsc], "FF": [ff]})
    buf = io.BytesIO()
    df.to_excel(buf, engine="openpyxl", index=False)
    buf.seek(0)
    return buf.getvalue()


def test_parse_and_workbook():
    data = make_excel(1.23, 4.56, 0.78)
    values, errors = asyncio.run(parse_excel(data))
    assert values == {"Voc": 1.23, "Jsc": 4.56, "FF": 0.78}
    assert not errors
    wb = create_workbook({("101", 1): values}, [date.today().isoformat()])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    loaded = load_workbook(buf)
    ws = loaded.active
    assert ws.cell(row=3, column=1).value == 1.23
    assert ws.cell(row=3, column=2).value == 4.56
    assert ws.cell(row=3, column=3).value == 0.78


def test_numeric_validation():
    data = make_excel("bad", 1, 2)
    values, errors = asyncio.run(parse_excel(data))
    assert values["Voc"] is None
    assert errors

def test_endpoints():
    client = TestClient(app)
    data = make_excel(0.5, 1.1, 0.9)
    files = {"files": ("101-1.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    res = client.post("/process", files=files)
    assert res.status_code == 200
    job_id = res.json()["job_id"]
    # wait for background task
    for _ in range(5):
        resp = client.get(f"/results/{job_id}")
        if resp.status_code == 200:
            break
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(row=3, column=1).value == 0.5
