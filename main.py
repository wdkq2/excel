from __future__ import annotations

import asyncio
import io
import uuid
from datetime import date
from typing import Dict, List, Tuple

import pandas as pd
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook

app = FastAPI()

jobs: Dict[str, Dict[str, object]] = {}


async def parse_excel(data: bytes) -> Tuple[Dict[str, float | None], List[str]]:
    """Extract Voc, Jsc and FF from an Excel file."""
    try:
        df = await asyncio.to_thread(
            pd.read_excel,
            io.BytesIO(data),
            engine="openpyxl",
            sheet_name=0,
            usecols=["Voc", "Jsc", "FF"],
        )
    except Exception as exc:  # pragma: no cover - pandas raises many variants
        return {"Voc": None, "Jsc": None, "FF": None}, [str(exc)]

    result: Dict[str, float | None] = {}
    errors: List[str] = []
    for col in ["Voc", "Jsc", "FF"]:
        value = None
        if col in df.columns and not df[col].dropna().empty:
            value = df[col].dropna().iloc[0]
        try:
            result[col] = float(value)
        except (TypeError, ValueError):
            result[col] = None
            errors.append(f"{col} invalid")
    return result, errors


def create_workbook(
    values: Dict[Tuple[str, int], Dict[str, float | None]], dates: List[str]
) -> Workbook:
    """Create output workbook from parsed values."""
    metrics = ["Voc", "Jsc", "FF"]
    wb = Workbook()
    ws = wb.active
    col = 1
    for metric in metrics:
        start_col = col
        for dt in dates:
            ws.cell(row=2, column=col, value=dt)
            col += 1
        end_col = col - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col, value=metric)

    device_rows: Dict[str, int] = {}
    base_row = 3
    for device, idx in sorted(values):
        if device not in device_rows:
            device_rows[device] = base_row
            base_row += 4
        row = device_rows[device] + idx - 1
        for m_i, metric in enumerate(metrics):
            for d_i, _ in enumerate(dates):
                col = m_i * len(dates) + d_i + 1
                ws.cell(row=row, column=col, value=values[(device, idx)].get(metric))
    return wb


async def process_files(job_id: str, uploaded: List[Tuple[str, bytes]]) -> None:
    """Background task to parse files and build result workbook."""
    jobs[job_id] = {"status": "processing"}
    values: Dict[Tuple[str, int], Dict[str, float | None]] = {}
    all_errors: List[str] = []
    for name, data in uploaded:
        try:
            device, idx_str = name.rsplit("-", 1)
            idx = int(idx_str.split(".")[0])
        except ValueError:
            all_errors.append(f"Invalid filename: {name}")
            continue
        metrics, errs = await parse_excel(data)
        values[(device, idx)] = metrics
        all_errors.extend(errs)
    wb = await asyncio.to_thread(
        create_workbook, values, [date.today().isoformat()]
    )
    buf = io.BytesIO()
    await asyncio.to_thread(wb.save, buf)
    buf.seek(0)
    jobs[job_id] = {"status": "ready", "buffer": buf, "errors": all_errors}


@app.post(
    "/process",
    summary="Upload multiple Excel files",
    response_description="Job identifier",
)
async def process(  # noqa: D401 - FastAPI handles docstring
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(
        ..., description="Excel files", openapi_extra={"examples": {"file": {"filename": "101-1.xlsx"}}}
    ),
):
    uploaded: List[Tuple[str, bytes]] = []
    for f in files:
        uploaded.append((f.filename, await f.read()))
    job_id = str(uuid.uuid4())
    background_tasks.add_task(process_files, job_id, uploaded)
    return {"job_id": job_id}


@app.get("/results/{job_id}")
async def get_results(job_id: str):
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Unknown job")
    if job["status"] != "ready":
        return JSONResponse({"status": "processing"})
    buf: io.BytesIO = job["buffer"]
    buf.seek(0)
    headers = {
        "Content-Disposition": f"attachment; filename=result_{job_id}.xlsx"
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
